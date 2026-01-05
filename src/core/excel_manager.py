import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import date
from typing import Dict, List, Any
import os
import shutil

class ExcelManager:
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.months_map = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
            5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
            9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }
    
    def create_report(self, schedule: Dict[date, List[Dict]], year: int, month: int, consultant_name: str = "Esteban Marulanda"):
        """
        Creates a new report file filled with the schedule.
        """
        month_name = self.months_map.get(month, "").upper()
        if not month_name:
            raise ValueError(f"Invalid month: {month}")

        # Create output filename
        output_filename = f"Reporte_{year}_{month_name}.xlsx"
        output_path = os.path.join(os.path.dirname(self.template_path), output_filename)
        
        # Copy template
        shutil.copyfile(self.template_path, output_path)
        
        workbook = openpyxl.load_workbook(output_path)
        
        if month_name not in workbook.sheetnames:
            print(f"Sheet {month_name} not found in template. Using active sheet.")
            sheet = workbook.active
        else:
            sheet = workbook[month_name]
            
        # Write Header Info if needed (Consultant Name)
        # Assuming row 2 is "Nombre del Consultor" based on image, maybe cell B2 or merged A2?
        # Image shows "Nombre del Consultor" in A2 maybe. Let's try to find it or just append rows.
        # The user said "llenar por fila una tarea".
        # Let's assume start row index = 4 (leaving header space).
        
        start_row = 4 # Guessing based on image (Row 1 headers, Row 2 Info, Row 3 headers maybe empty?)
        
        # Yellow Fill style
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        current_row = start_row
        
        # Sort days to write in order
        sorted_days = sorted(schedule.keys())
        
        for day in sorted_days:
            tasks = schedule[day]
            day_num = day.day
            
            for task in tasks:
                # Columns:
                # A: Task Name
                # B: Client/Project
                # C: Hours
                # D-AH: Days 1-31.
                # '1' is in Col D (4th col). So Day X is at Col index 3 + X.
                
                sheet.cell(row=current_row, column=1, value=task.get('task_name', ''))
                sheet.cell(row=current_row, column=2, value=task.get('client_project', ''))
                sheet.cell(row=current_row, column=3, value=task.get('hours', 0))
                
                # Mark the day column
                day_col_idx = 3 + day_num
                cell = sheet.cell(row=current_row, column=day_col_idx)
                cell.value = "X" # Or input hours? "Then paint yellow the day". Usually just X or hours.
                # User said: "Luego pintar en amarillo el dia que se hizo"
                cell.fill = yellow_fill
                
                current_row += 1
                
        workbook.save(output_path)
        return output_path
